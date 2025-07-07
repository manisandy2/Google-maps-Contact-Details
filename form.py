import time

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from openpyxl import Workbook
import datetime

date = datetime.datetime.now().strftime("%d-%m-%Y")

web_url = r"https://www.google.com/maps"
extension = ".xlsx"


wb = Workbook()
ws = wb.active


class GoogleMaps:

    def __init__(self,Title,Location,Driver_Path,Save_Excel):
        self.Title = Title
        self.Location = Location
        self.Driver = webdriver.Chrome(executable_path=Driver_Path) 
        self.Location_save = Save_Excel
        
      

    def browser_run(self):
        self.Driver.get(web_url)
        # driver.get(web_url)
        

    def input_fields(self,  **kwargs):
        self.Driver.find_element(By.ID, "searchboxinput").send_keys(kwargs.get("find"))
        time.sleep(3)

    def input_fields_post(self):
        self.Driver.find_element(By.ID, "searchbox-searchbutton").click()
        time.sleep(3)

    def scroll(self):
        for r in range(1, 10):
            print(r)
            self.Driver.find_element(By.CLASS_NAME, "hfpxzc").send_keys(Keys.PAGE_DOWN)
            time.sleep(2)

    def excel_link_title(self):
        ws.cell(row=1,column=1).value = "Link"

    def excel_title(self):
        ws.cell(row=1,column=1).value = "Link"
        ws.cell(row=1,column=2).value = "Name"
        ws.cell(row=1,column=3).value = "Address"
        ws.cell(row=1,column=4).value = "Website"
        ws.cell(row=1,column=5).value = "Phone"

    def get_link(self, r=2):
        for link in self.Driver.find_elements(By.CLASS_NAME, "hfpxzc"):
            print(link.get_attribute("href"))
            ws.cell(row=r, column=1).value = link.get_attribute("href")
            r = r + 1
        self.save_excel(title=self.Title + " in " + self.Location  +" data ")

    def save_excel(self,**kwargs):
        wb.save(self.Location_save[:-1] +  kwargs.get("title") + date + extension)


    def search_text(self):
        print(self.Title,self.Location)
        self.browser_run()
        self.input_fields(find=self.Title + " in " + self.Location)
        self.input_fields_post()
        self.scroll()
        self.excel_link_title()
        self.get_link()
        time.sleep(3)
        self.get_data()

    def get_phone(self):
        try:
            ph = self.Driver.find_element(By.XPATH, '//*[@id="QA0Szd"]/div/div/div[1]/div[2]/div/div[1]/div/div/div[7]/div[6]/button/div/div[2]/div[1]')
            return ph.text
        except:
            pass
    def element(self,**kwargs):
        try:
            name = self.Driver.find_element(kwargs.get("by"), kwargs.get("value")).text
            self.element_print(heading=kwargs.get("heading"), value=name)
            return name
        except:
            return None

    def element_print(self,**kwargs):
        print(kwargs.get("heading"), ":", kwargs.get("value"))

    def maps_data_range(self):
        self.excel_title()
        for r in range(2, 130): #130
            print(r)

            if ws.cell(row=r, column=1).value:
                self.Driver.get(ws.cell(row=r, column=1).value)
                name = self.element(heading="Name", by=By.TAG_NAME, value="h1")
                address = self.element(heading="Address", by=By.CLASS_NAME, value="Io6YTe")
                website = self.element(heading="Website", by=By.CLASS_NAME, value="ITvuef")
                phone = self.get_phone()

                print("Phone :", phone)

                ws.cell(row=r, column=2).value = name
                ws.cell(row=r, column=3).value = address
                ws.cell(row=r, column=4).value = website
                ws.cell(row=r, column=5).value = phone


    def get_data(self):
        print(self.Location_save[:-1] +  self.Title + " in " + self.Location  +" data " + date + extension)
        self.maps_data_range()
        self.save_excel(title=self.Title + " in " + self.Location  +" data_details ")




