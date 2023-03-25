from selenium import webdriver
from openpyxl import Workbook,load_workbook
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
import datetime
search_data = "Used appliances Dealers in Karnataka"


date = datetime.datetime.now().strftime("%d-%m-%Y")
driver_path = r"E:\Python project\Google maps contact details\Driver\chromedriver_win32\chromedriver.exe"

wb = load_workbook(r"E:\Python project\Google maps contact details\Excel\ Used appliances Dealers in Karnataka  24-03-2023.xlsx")
ws = wb.active

driver = webdriver.Chrome(executable_path=driver_path)

for r in range(2,130):
    print(r)
    name = ""
    address = ""
    website = ""
    phone = ""

    driver.get(ws.cell(row=r,column=1).value)
    name = driver.find_element(By.TAG_NAME,"h1").text
    address = driver.find_element(By.CLASS_NAME,"Io6YTe").text
    try:
        website = driver.find_element(By.CLASS_NAME,"ITvuef").text
    except:

        None

    for ph in driver.find_elements(By.TAG_NAME,"button"):
        if len(ph.text) == 12:
            # print(ph.text)
            phone = ph.text

    print("Name :",name)
    print("Address :",address)
    print("WebSite :", website)
    print("Phone :",phone)
    ws.cell(row=r,column=2).value = name
    ws.cell(row=r,column=3).value = address
    ws.cell(row=r,column=4).value = website
    ws.cell(row=r,column=5).value = phone

    wb.save(r"E:\Python project\Google maps contact details\Excel\details " + search_data + " " + date + ".xlsx")


# for r in l:
#     print(driver.find_element(By.CLASS_NAME, r).text)


# for r in l:
#     print(r)
# Phone = driver.find_element(By.TAG_NAME,"button").get_attribute("class")

# Io6YTe fontBodyMedium
# rogA2c ITvuef
# RcCsl fVHpi w4vB1d NOE9ve M0S7ae AG25L